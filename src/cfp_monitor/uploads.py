"""URL extraction and scoped row-context loading for customer list uploads."""
from __future__ import annotations

import io
import re
from datetime import date, datetime

from .scoring import normalize_url

URL_RE = re.compile(r'https?://[^\s"\'<>)\]]+')


def _urls_in_text(value: object) -> list[str]:
    return URL_RE.findall(str(value or ""))


def _context_value(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return str(value or "").strip()


def _row_context(row: tuple[object, ...], industry: object = None) -> dict | None:
    """The narrowly scoped context used for one-hop aggregator resolution.

    `industry` (from an optional Industry column) is attached as input metadata when present;
    it is NOT used for resolution, only carried through for filtering downstream."""
    context = {
        "name": _context_value(row[0]) if len(row) > 0 else "",
        "location": _context_value(row[2]) if len(row) > 2 else "",
        "dates": _context_value(row[3]) if len(row) > 3 else "",
    }
    ind = _context_value(industry)
    if ind:
        context["industry"] = ind
    return context if any(context.values()) else None


def _find_industry_col(sheet) -> int | None:
    """1-based index of a header cell literally named 'Industry' (case-insensitive), if any.
    Scans only the first few rows so a stray data cell can't be mistaken for the header."""
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        for idx, val in enumerate(row, start=1):
            if isinstance(val, str) and val.strip().lower() == "industry":
                return idx
    return None


# Header text -> role. The FIRST alias that matches (in listed order) claims the role, so a
# "CONFERENCE URL" is preferred over a "SUBMISSION URL" when a sheet has both (conference lists
# do; award lists only have SUBMISSION URL). This is what lets one intake read both layouts:
# conferences crawl the event site, awards crawl the award/submission page.
_HEADER_ROLES = {
    "url": ("conference url", "submission url", "award url", "url", "link"),
    "name": ("conference", "award", "event name", "event", "name"),
    "location": ("location", "venue", "city"),
    "dates": ("start dates", "event dates", "conference dates", "dates"),
    "industry": ("industry", "market"),
}


def _norm_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _column_map(sheet) -> dict[str, int]:
    """Map role -> 1-based column index from the header row. Empty if no headers are recognized
    (which triggers the legacy position-based fallback). ONLY the 'url' column ever becomes a
    crawl target, so other columns that happen to hold links (CATEGORIES, COORDINATOR EMAIL) are
    never crawled -- the same safety property as the old Column-B-only rule."""
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        headers = {_norm_header(v): i + 1 for i, v in enumerate(row) if v not in (None, "")}
        roles: dict[str, int] = {}
        for role, aliases in _HEADER_ROLES.items():
            for alias in aliases:
                if alias in headers:
                    roles[role] = headers[alias]
                    break
        if "url" in roles:
            # Robustness for the classic conference layout (name=A, url=B): if a sheet omits the
            # LOCATION / START DATES headers, still read them from C/D as the old position rule
            # did. This can NEVER fire for award sheets (their URL is far right, not column B),
            # so it can't mislabel award date columns as location/dates.
            if roles["url"] == 2 and roles.get("name") == 1:
                roles.setdefault("location", 3)
                roles.setdefault("dates", 4)
            return roles
    return {}


def _context_from_roles(row: tuple, roles: dict[str, int]) -> dict | None:
    """Row context from headed columns only. Absent roles stay absent -- so an award sheet
    (no LOCATION/START DATES columns) never mislabels its date columns as location/dates."""
    def cell(role: str) -> object:
        ci = roles.get(role)
        return row[ci - 1] if ci and len(row) >= ci else None

    context = {
        "name": _context_value(cell("name")),
        "location": _context_value(cell("location")),
        "dates": _context_value(cell("dates")),
    }
    ind = _context_value(cell("industry"))
    if ind:
        context["industry"] = ind
    return context if any(context.values()) else None


def _xlsx_urls_and_contexts(data: bytes) -> tuple[list[str], list[dict | None]]:
    """Read crawl-target URLs + scoped row context from a workbook, by column HEADER.

    The URL column is chosen by header ('CONFERENCE URL' for event lists, 'SUBMISSION URL' for
    award lists), and name/location/dates are read only from columns that carry those headers.
    Only the designated URL column becomes a crawl target; hyperlinks, workbook XML, notes and
    every other column are ignored. If no headers are recognized, fall back to the legacy
    position rule (Column B URL, A/C/D context, optional Industry column).
    """
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    except Exception:
        return [], []

    try:
        urls: list[str] = []
        contexts: list[dict | None] = []
        for sheet in workbook.worksheets:
            roles = _column_map(sheet)
            if roles:
                url_col = roles["url"]
                max_col = max(roles.values())
                for row in sheet.iter_rows(min_col=1, max_col=max_col, values_only=True):
                    context = _context_from_roles(row, roles)
                    cell = row[url_col - 1] if len(row) >= url_col else None
                    for url in _urls_in_text(cell):
                        urls.append(url)
                        contexts.append(context)
            else:
                industry_col = _find_industry_col(sheet)
                max_col = max(4, industry_col or 0)
                for row in sheet.iter_rows(min_col=1, max_col=max_col, values_only=True):
                    ind = row[industry_col - 1] if (industry_col and len(row) >= industry_col) else None
                    context = _row_context(row, ind)
                    for url in _urls_in_text(row[1]):
                        urls.append(url)
                        contexts.append(context)
        return urls, contexts
    finally:
        workbook.close()


def normalize_urls_and_contexts_audited(
    raw_urls: list[str], raw_contexts: list[dict | None]
) -> tuple[list[str], list[dict | None], dict]:
    """Normalize/dedupe URLs, keep the first useful row context aligned, AND return an audit
    manifest explaining every drop — so a "54 rows in -> 51 targets" result is fully traceable.

    The manifest is: {raw_count, kept_count, dropped_count, dropped:[{url, reason, duplicate_of}]}
    where reason is 'not_a_url' or 'duplicate' (duplicate_of names the surviving URL it folded into).
    """
    seen: dict[str, int] = {}
    urls: list[str] = []
    contexts: list[dict | None] = []
    dropped: list[dict] = []
    for index, value in enumerate(raw_urls):
        original = (value or "").strip()
        url = original.rstrip(",;")
        if not url.lower().startswith("http"):
            dropped.append({"url": original, "reason": "not_a_url", "duplicate_of": None})
            continue
        context = raw_contexts[index] if index < len(raw_contexts) else None
        key = normalize_url(url)
        if key in seen:
            existing = seen[key]
            if context and not contexts[existing]:
                contexts[existing] = context
            dropped.append({"url": url, "reason": "duplicate", "duplicate_of": urls[existing]})
            continue
        seen[key] = len(urls)
        urls.append(url)
        contexts.append(context)
    manifest = {
        "raw_count": len(raw_urls),
        "kept_count": len(urls),
        "dropped_count": len(dropped),
        "dropped": dropped,
    }
    return urls, contexts, manifest


def normalize_urls_and_contexts(raw_urls: list[str], raw_contexts: list[dict | None]) -> tuple[list[str], list[dict | None]]:
    """Normalize/dedupe URLs while keeping the first useful row context aligned."""
    urls, contexts, _ = normalize_urls_and_contexts_audited(raw_urls, raw_contexts)
    return urls, contexts


def uploaded_urls_and_contexts(name: str, data: bytes) -> tuple[list[str], list[dict | None]]:
    """Load uploaded URLs and per-URL context, preserving list-index alignment."""
    if name.lower().endswith(".xlsx"):
        return _xlsx_urls_and_contexts(data)
    urls = _urls_in_text(data.decode("utf-8-sig", "ignore"))
    return urls, [None] * len(urls)


def urls_from_upload(name: str, data: bytes) -> list[str]:
    """Extract every HTTP(S) URL from a .txt, .csv, or .xlsx list upload."""
    return uploaded_urls_and_contexts(name, data)[0]
