"""Write an ANNOTATED copy of the customer's master .xlsx: their exact sheet, with our
differences highlighted (cell fill) + a comment carrying our value and the source, plus a
summary tab. Uses openpyxl (offline, no auth). The reconciliation logic itself is in
`reconcile.py` (pure/tested); this module is the presentation layer.
"""
from __future__ import annotations

from datetime import date

from .reconcile import (
    reconcile_all, CONFIRMED, CHANGED, GAP_FILLED, UNVERIFIED, NOT_CRAWLED, FIELDS,
)
from .storage import Store, normalize_key

# Category -> (fill hex, human label). Semantic, not the app accent.
_STYLE = {
    GAP_FILLED:  ("C6EFCE", "Gap filled — we found a value your cell was missing"),
    CHANGED:     ("FFEB9C", "Changed — our value differs from yours; please review"),
    CONFIRMED:   ("EAF3EA", "Confirmed — our crawl matches your sheet"),
    UNVERIFIED:  ("E2E2E2", "Unverified — we couldn't confirm this row this run"),
    NOT_CRAWLED: ("F0E0E0", "Not crawled — this URL wasn't in the run"),
}


def _our_records(db_path: str) -> dict:
    """normalized-url -> our fact dict (with quality + last_checked) from the source of truth."""
    store = Store(db_path)
    try:
        out = {}
        for r in store.all_records():
            out[r["key"]] = {
                "name": r["name"], "location": r["location"], "start_dates": r["conference_dates"],
                "submission_deadline": r["cfp_close_date"], "submission_url": r["submission_url"],
                "status": r["cfp_status"], "url": r["url"], "last_checked": r["last_checked"],
                "quality": r["quality"],
            }
        return out
    finally:
        store.close()


def annotate_workbook(src_xlsx: str, dst_xlsx: str, our_by_key: dict) -> dict:
    """Annotate `src_xlsx` -> `dst_xlsx`. Returns the summary counts. Pure I/O around reconcile."""
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill, Font

    wb = load_workbook(src_xlsx)
    ws = wb.active

    # Map header text -> column index from row 1 (robust to column reordering).
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip().upper()] = c
    url_col = headers.get("CONFERENCE URL")
    if not url_col:
        raise ValueError("Could not find a 'CONFERENCE URL' column in the sheet.")

    # Build customer-row dicts (header -> value). Skip blank rows — a row with no CONFERENCE URL
    # is empty spreadsheet padding, not a conference we failed to crawl.
    customer_rows = []
    for row in range(2, ws.max_row + 1):
        url = ws.cell(row=row, column=url_col).value
        if not (str(url or "").strip().lower().startswith("http")):
            continue
        rec = {h: (ws.cell(row=row, column=col).value or "") for h, col in headers.items()}
        rec["_row"] = row
        customer_rows.append(rec)

    recons, counts = reconcile_all(customer_rows, our_by_key, normalize_key)

    fills = {cat: PatternFill(fill_type="solid", fgColor=hexc) for cat, (hexc, _) in _STYLE.items()}
    for tr, rr in zip(customer_rows, recons):
        row = tr["_row"]
        src = f"\nSource: {rr.source_url}" if rr.source_url else ""
        chk = f" (checked {rr.checked[:10]})" if rr.checked else ""
        # Row-level flags land on the URL cell.
        if rr.row_category in (UNVERIFIED, NOT_CRAWLED):
            cell = ws.cell(row=row, column=url_col)
            cell.fill = fills[rr.row_category]
            cell.comment = Comment(_STYLE[rr.row_category][1] + chk + src, "cfp-monitor")
        # Field-level annotations.
        for f in rr.fields:
            if f.category == CONFIRMED:
                continue    # keep the sheet quiet where we agree
            col = headers.get(f.column.upper())
            if not col:
                continue
            cell = ws.cell(row=row, column=col)
            cell.fill = fills[f.category]
            if f.category == GAP_FILLED:
                body = f"We found: {f.ours}{chk}{src}"
            else:  # CHANGED
                body = f"Ours: {f.ours}\nYours: {f.theirs}\nPlease review.{chk}{src}"
            cell.comment = Comment(body, "cfp-monitor")

    _write_summary(wb, counts, len(customer_rows))
    wb.save(dst_xlsx)
    return counts


def _write_summary(wb, counts: dict, n_rows: int) -> None:
    from openpyxl.styles import PatternFill, Font

    ws = wb.create_sheet("Reconciliation", 0)
    ws["A1"] = "Reconciliation summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated {date.today().isoformat()} · {n_rows} rows compared"
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A4"] = "Category"; ws["B4"] = "Count"; ws["C4"] = "Meaning"
    for c in ("A4", "B4", "C4"):
        ws[c].font = Font(bold=True)
    order = [CHANGED, GAP_FILLED, UNVERIFIED, NOT_CRAWLED, CONFIRMED]
    r = 5
    for cat in order:
        hexc, label = _STYLE[cat]
        ws.cell(row=r, column=1, value=cat.replace("_", " ").title()).fill = \
            PatternFill(fill_type="solid", fgColor=hexc)
        ws.cell(row=r, column=2, value=counts.get(cat, 0))
        ws.cell(row=r, column=3, value=label)
        r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 64


def annotate_from_db(src_xlsx: str, dst_xlsx: str, db_path: str = "cfp_monitor.db") -> dict:
    """Convenience: reconcile the customer sheet against the source-of-truth DB."""
    return annotate_workbook(src_xlsx, dst_xlsx, _our_records(db_path))
