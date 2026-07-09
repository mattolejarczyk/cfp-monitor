"""Offline tests for the reconciliation taxonomy (pure logic; no xlsx needed)."""
from cfp_monitor.reconcile import (
    reconcile_row, reconcile_all, CONFIRMED, CHANGED, GAP_FILLED, UNVERIFIED, NOT_CRAWLED,
)
from cfp_monitor.storage import normalize_key


def _their(**kw):
    base = {"CONFERENCE": "", "CONFERENCE URL": "https://c.com", "LOCATION": "",
            "START DATES": "", "SUBMISSION DEADLINE": "", "SUBMISSION URL": ""}
    base.update(kw)
    return base


def test_gap_filled_and_confirmed_and_changed():
    their = _their(CONFERENCE="Alpha", LOCATION="", **{"SUBMISSION DEADLINE": "June 1, 2026"})
    our = {"name": "Alpha", "location": "Berlin, Germany", "submission_deadline": "May 15, 2026",
           "url": "https://c.com", "last_checked": "2026-07-07", "quality": "PASS"}
    rr = reconcile_row(their, our)
    cats = {f.column: f.category for f in rr.fields}
    assert cats["CONFERENCE"] == CONFIRMED
    assert cats["LOCATION"] == GAP_FILLED                  # theirs blank, we found it
    assert cats["SUBMISSION DEADLINE"] == CHANGED          # both present, differ
    assert rr.row_category is None


def test_no_annotation_when_we_add_nothing():
    their = _their(CONFERENCE="Alpha", LOCATION="Paris")
    our = {"name": "Alpha", "location": "", "url": "https://c.com", "quality": "PASS"}
    rr = reconcile_row(their, our)
    cols = {f.column for f in rr.fields}
    assert "LOCATION" not in cols                          # theirs present, ours blank -> untouched
    assert {f.column for f in rr.changed} == set()


def test_unverified_row():
    our = {"name": "", "url": "https://c.com", "quality": "ERROR"}
    rr = reconcile_row(_their(CONFERENCE="Alpha"), our)
    assert rr.row_category == UNVERIFIED


def test_not_crawled_row():
    rr = reconcile_row(_their(CONFERENCE="Alpha"), None)
    assert rr.row_category == NOT_CRAWLED


def test_reconcile_all_counts():
    rows = [_their(CONFERENCE="Alpha"), _their(**{"CONFERENCE URL": "https://d.com", "CONFERENCE": "Beta"})]
    our_by_key = {
        normalize_key("https://c.com"): {"name": "Alpha", "location": "Berlin", "url": "https://c.com", "quality": "PASS"},
        # d.com has no record -> NOT_CRAWLED
    }
    _, counts = reconcile_all(rows, our_by_key, normalize_key)
    assert counts[GAP_FILLED] == 1      # c.com location filled
    assert counts[NOT_CRAWLED] == 1     # d.com


def test_date_aware_confirm():
    # Their Excel serial-date string vs our verbatim text -> same year+month -> CONFIRMED, not CHANGED.
    their = _their(CONFERENCE="Alpha", **{"START DATES": "2026-10-13 00:00:00"})
    our = {"name": "Alpha", "start_dates": "13-15 October 2026", "url": "https://c.com", "quality": "PASS"}
    cats = {f.column: f.category for f in reconcile_row(their, our).fields}
    assert cats["START DATES"] == CONFIRMED


def test_xlsx_writer_roundtrip(tmp_path=None):
    import tempfile, os
    from openpyxl import Workbook, load_workbook
    from cfp_monitor.reconcile_xlsx import annotate_workbook
    from cfp_monitor.storage import normalize_key

    d = tempfile.mkdtemp()
    src = os.path.join(d, "src.xlsx"); dst = os.path.join(d, "out.xlsx")
    wb = Workbook(); ws = wb.active; ws.title = "Conference List"
    ws.append(["CONFERENCE", "CONFERENCE URL", "LOCATION", "START DATES", "SUBMISSION DEADLINE", "SUBMISSION URL"])
    ws.append(["Alpha", "https://c.com", "", "", "", ""])            # location gap to fill
    ws.append(["", "", "", "", "", ""])                              # blank padding row -> skipped
    wb.save(src)

    our = {normalize_key("https://c.com"): {"name": "Alpha", "location": "Berlin, Germany",
           "start_dates": "", "submission_deadline": "", "submission_url": "",
           "url": "https://c.com", "last_checked": "2026-07-07", "quality": "PASS"}}
    counts = annotate_workbook(src, dst, our)
    assert counts[GAP_FILLED] == 1 and counts[NOT_CRAWLED] == 0     # blank row ignored

    wb2 = load_workbook(dst)
    assert "Reconciliation" in wb2.sheetnames
    ws2 = wb2["Conference List"]
    loc = ws2.cell(row=2, column=3)                                 # LOCATION cell for Alpha
    assert loc.fill.fgColor.rgb.endswith("C6EFCE")                  # gap-filled green
    assert loc.comment and "Berlin" in loc.comment.text


def _run():
    import sys
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    bad = 0
    for fn in fns:
        try:
            fn(); print(f"PASS {fn.__name__}")
        except Exception as e:
            bad += 1; print(f"FAIL {fn.__name__}: {e!r}")
    print(f"--- {len(fns)-bad}/{len(fns)} passed ---")
    sys.exit(1 if bad else 0)


if __name__ == "__main__":
    _run()
