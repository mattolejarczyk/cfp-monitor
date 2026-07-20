"""Regression tests for customer URL-list uploads."""
import io
from datetime import date

from openpyxl import Workbook

from cfp_monitor.uploads import (
    normalize_urls_and_contexts,
    normalize_urls_and_contexts_audited,
    uploaded_urls_and_contexts,
    urls_from_upload,
)


def _xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Conference"
    sheet["B1"] = "Conference URL"
    sheet["A2"] = "IROS 2026"
    sheet["B2"] = "https://iros2026.org/"
    sheet["C2"] = "Kyoto, Japan"
    sheet["D2"] = date(2026, 9, 27)
    sheet["A3"] = "RoboBusiness"
    sheet["B3"] = "https://www.robobusiness.com/"
    sheet["C3"] = "Santa Clara, California"
    sheet["D3"] = "2026-10-14"
    out = io.BytesIO()
    workbook.save(out)
    workbook.close()
    return out.getvalue()


def test_xlsx_reads_visible_cells_not_internal_xml_namespaces():
    urls = urls_from_upload("two-conferences.xlsx", _xlsx_bytes())
    assert urls == [
        "https://iros2026.org/",
        "https://www.robobusiness.com/",
    ]
    assert all("openxmlformats" not in url for url in urls)


def test_xlsx_keeps_scoped_row_context_aligned_to_each_column_b_url():
    urls, contexts = uploaded_urls_and_contexts("two-conferences.xlsx", _xlsx_bytes())
    assert urls == ["https://iros2026.org/", "https://www.robobusiness.com/"]
    assert contexts == [
        {"name": "IROS 2026", "location": "Kyoto, Japan", "dates": "2026-09-27"},
        {"name": "RoboBusiness", "location": "Santa Clara, California", "dates": "2026-10-14"},
    ]


def test_normalization_dedupes_urls_without_losing_their_aligned_context():
    urls, contexts = normalize_urls_and_contexts(
        [" https://iros2026.org/ ", "https://iros2026.org", "https://robobusiness.com"],
        [{"name": "IROS 2026"}, None, {"name": "RoboBusiness"}],
    )
    assert urls == ["https://iros2026.org/", "https://robobusiness.com"]
    assert contexts == [{"name": "IROS 2026"}, {"name": "RoboBusiness"}]


def test_text_upload_still_extracts_urls():
    urls = urls_from_upload("list.txt", b"IROS https://iros2026.org/\n")
    assert urls == ["https://iros2026.org/"]


def test_audit_manifest_explains_every_drop():
    urls, contexts, manifest = normalize_urls_and_contexts_audited(
        ["https://a.org/", "https://a.org", "not-a-link", "https://b.org"],
        [None, None, None, None],
    )
    assert urls == ["https://a.org/", "https://b.org"]
    assert manifest["raw_count"] == 4
    assert manifest["kept_count"] == 2
    assert manifest["dropped_count"] == 2
    reasons = {d["reason"] for d in manifest["dropped"]}
    assert reasons == {"duplicate", "not_a_url"}
    dupe = next(d for d in manifest["dropped"] if d["reason"] == "duplicate")
    assert dupe["duplicate_of"] == "https://a.org/"   # names the survivor it folded into


def test_optional_industry_column_is_read_per_row():
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "Conference"
    sheet["B1"] = "Conference URL"
    sheet["E1"] = "Industry"          # anywhere; matched by header name
    sheet["A2"] = "IROS 2026"
    sheet["B2"] = "https://iros2026.org/"
    sheet["E2"] = "Robotics"
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    urls, contexts = uploaded_urls_and_contexts("with-industry.xlsx", out.getvalue())
    assert urls == ["https://iros2026.org/"]
    assert contexts[0]["industry"] == "Robotics"


def test_no_industry_column_leaves_context_industry_absent():
    urls, contexts = uploaded_urls_and_contexts("two-conferences.xlsx", _xlsx_bytes())
    assert all("industry" not in (c or {}) for c in contexts)


def _award_xlsx_bytes() -> bytes:
    """An award workbook: no CONFERENCE URL / LOCATION / START DATES; URL lives in a
    'SUBMISSION URL' column further right, and C/D are dates (not location/dates)."""
    wb = Workbook()
    ws = wb.active
    for col, head in enumerate(
        ["AWARD", "OVERVIEW", "LATEST UPDATE", "SUBMISSION DEADLINE", "SUBMISSION DATE VERIFIED",
         "SUBMISSION STATUS", "STATUS DETAILS", "SUBMISSION URL", "COORDINATOR EMAIL", "CATEGORIES"],
        start=1):
        ws.cell(row=1, column=col, value=head)
    ws.cell(row=2, column=1, value="ASTORS Homeland Security Awards")
    ws.cell(row=2, column=2, value="Security awards program")
    ws.cell(row=2, column=4, value=date(2026, 6, 15))            # a DATE in col D
    ws.cell(row=2, column=8, value="https://americansecuritytoday.com/astors-awards/")   # URL in H
    ws.cell(row=2, column=10, value="https://example.com/a-category-link")               # must NOT be crawled
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def test_award_layout_reads_submission_url_column_by_header():
    urls, contexts = uploaded_urls_and_contexts("Utility Global Award List.xlsx", _award_xlsx_bytes())
    # only the SUBMISSION URL column is a crawl target -- the CATEGORIES link is ignored.
    assert urls == ["https://americansecuritytoday.com/astors-awards/"]
    # name comes from AWARD; the date columns are NOT mislabeled as location/dates.
    assert contexts[0]["name"] == "ASTORS Homeland Security Awards"
    assert not contexts[0].get("location")
    assert not contexts[0].get("dates")


def test_conference_layout_still_reads_column_b_and_full_context():
    # Regression: header-driven intake must leave conference behavior byte-identical.
    urls, contexts = uploaded_urls_and_contexts("two-conferences.xlsx", _xlsx_bytes())
    assert urls == ["https://iros2026.org/", "https://www.robobusiness.com/"]
    assert contexts[0] == {"name": "IROS 2026", "location": "Kyoto, Japan", "dates": "2026-09-27"}


def _run():
    import sys
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    failures = 0
    for fn in fns:
        try:
            fn()
            print(f"PASS {fn.__name__}")
        except Exception as exc:
            failures += 1
            print(f"FAIL {fn.__name__}: {exc!r}")
    print(f"--- {len(fns) - failures}/{len(fns)} passed ---")
    sys.exit(1 if failures else 0)


if __name__ == "__main__":
    _run()
